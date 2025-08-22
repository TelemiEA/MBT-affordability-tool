#!/usr/bin/env python3
"""
Affordability Benchmarking Tool for Mortgage Lenders

This tool compares mortgage lending amounts across multiple lenders
for different scenarios (Vanilla and Self-Employed applicants).
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple, Optional
import os
from datetime import datetime


class AffordabilityTool:
    """Main class for the affordability benchmarking tool."""
    
    def __init__(self):
        self.lenders = [
            "Gen H", "Accord", "Skipton", "Kensington", "Precise", "Atom",
            "Clydesdale", "Newcastle", "Metro", "Nottingham", "Hinckley & Rugby",
            "Leeds", "Principality", "Coventry", "Santander"
        ]
        
        self.scenarios = {
            'vanilla': self._create_vanilla_scenarios(),
            'self_employed': self._create_self_employed_scenarios()
        }
        
        # Storage for borrowing amounts
        self.borrowing_data = {}
    
    def _create_vanilla_scenarios(self) -> List[Dict]:
        """Create hardcoded vanilla scenarios."""
        scenarios = []
        
        # Joint scenarios
        joint_incomes = [30000, 35000, 40000, 45000, 50000, 60000, 70000, 80000]
        for income in joint_incomes:
            scenarios.append({
                'id': f'joint_vanilla_{income}',
                'applicant_type': 'Joint',
                'applicant1_income': income,
                'applicant2_income': income,
                'applicant1_employment': 'Employed',
                'applicant2_employment': 'Employed',
                'age': 30,
                'term': 35,
                'notes': 'Standard employed applicants'
            })
        
        # Single scenarios
        single_incomes = [30000, 35000, 40000, 45000, 50000, 60000, 70000, 80000]
        for income in single_incomes:
            scenarios.append({
                'id': f'single_vanilla_{income}',
                'applicant_type': 'Single',
                'applicant1_income': income,
                'applicant2_income': None,
                'applicant1_employment': 'Employed',
                'applicant2_employment': None,
                'age': 30,
                'term': 35,
                'notes': 'Standard employed applicant'
            })
        
        return scenarios
    
    def _create_self_employed_scenarios(self) -> List[Dict]:
        """Create hardcoded self-employed scenarios."""
        scenarios = []
        
        # Joint scenarios (one self-employed, one employed)
        joint_incomes = [30000, 35000, 40000, 45000, 50000, 60000, 70000, 80000]
        for income in joint_incomes:
            scenarios.append({
                'id': f'joint_self_employed_{income}',
                'applicant_type': 'Joint',
                'applicant1_income': income,
                'applicant2_income': income,
                'applicant1_employment': 'Self-Employed',
                'applicant2_employment': 'Employed',
                'age': 30,
                'term': 35,
                'notes': 'Self-employed: Latest year double previous year'
            })
        
        # Single self-employed scenarios
        single_incomes = [30000, 35000, 40000, 45000, 50000, 60000, 70000, 80000]
        for income in single_incomes:
            scenarios.append({
                'id': f'single_self_employed_{income}',
                'applicant_type': 'Single',
                'applicant1_income': income,
                'applicant2_income': None,
                'applicant1_employment': 'Self-Employed',
                'applicant2_employment': None,
                'age': 30,
                'term': 35,
                'notes': 'Self-employed: Latest year double previous year'
            })
        
        return scenarios
    
    def input_borrowing_amounts(self, scenario_type: str, scenario_id: str, 
                              amounts: Dict[str, float]) -> None:
        """
        Input borrowing amounts for a specific scenario.
        
        Args:
            scenario_type: 'vanilla' or 'self_employed'
            scenario_id: Unique identifier for the scenario
            amounts: Dictionary mapping lender names to borrowing amounts
        """
        if scenario_type not in self.borrowing_data:
            self.borrowing_data[scenario_type] = {}
        
        self.borrowing_data[scenario_type][scenario_id] = amounts
    
    def calculate_statistics(self, amounts: Dict[str, float]) -> Dict[str, float]:
        """
        Calculate average, Gen H difference, and rank for a scenario.
        
        Args:
            amounts: Dictionary of lender amounts
            
        Returns:
            Dictionary with calculated statistics
        """
        # Remove None values and convert to float
        valid_amounts = {k: float(v) for k, v in amounts.items() if v is not None and v != ''}
        
        if not valid_amounts:
            return {'average': 0, 'gen_h_diff': 0, 'gen_h_rank': 0}
        
        # Calculate average
        average = np.mean(list(valid_amounts.values()))
        
        # Gen H specific calculations
        gen_h_amount = valid_amounts.get('Gen H', 0)
        gen_h_diff = gen_h_amount - average
        
        # Calculate rank (1 = highest amount)
        sorted_amounts = sorted(valid_amounts.values(), reverse=True)
        gen_h_rank = sorted_amounts.index(gen_h_amount) + 1 if gen_h_amount in sorted_amounts else len(sorted_amounts)
        
        return {
            'average': average,
            'gen_h_diff': gen_h_diff,
            'gen_h_rank': gen_h_rank
        }
    
    def create_results_dataframe(self, scenario_type: str) -> pd.DataFrame:
        """
        Create a DataFrame with results for a specific scenario type.
        
        Args:
            scenario_type: 'vanilla' or 'self_employed'
            
        Returns:
            Formatted DataFrame with results
        """
        if scenario_type not in self.borrowing_data:
            return pd.DataFrame()
        
        scenarios = self.scenarios[scenario_type]
        data = []
        
        for scenario in scenarios:
            scenario_id = scenario['id']
            if scenario_id not in self.borrowing_data[scenario_type]:
                continue
            
            amounts = self.borrowing_data[scenario_type][scenario_id]
            stats = self.calculate_statistics(amounts)
            
            # Create row data
            row = {
                'Scenario': f"{scenario['applicant_type']} - £{scenario['applicant1_income']:,}",
                'Applicant Type': scenario['applicant_type']
            }
            
            # Add lender amounts
            for lender in self.lenders:
                row[lender] = amounts.get(lender, 0)
            
            # Add calculated statistics
            row['Average'] = stats['average']
            row['Gen H Difference'] = stats['gen_h_diff']
            row['Gen H Rank'] = stats['gen_h_rank']
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def export_to_excel(self, filename: str = None) -> str:
        """
        Export results to Excel file with proper formatting.
        
        Args:
            filename: Optional filename, defaults to timestamped name
            
        Returns:
            Path to created Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"affordability_benchmarking_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each scenario type
        for scenario_type in ['vanilla', 'self_employed']:
            df = self.create_results_dataframe(scenario_type)
            if df.empty:
                continue
            
            sheet_name = "Vanilla" if scenario_type == 'vanilla' else "Self-Employed"
            ws = wb.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            self._format_worksheet(ws, df)
        
        # Save workbook
        wb.save(filename)
        return filename
    
    def _format_worksheet(self, ws, df):
        """Apply formatting to Excel worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Format data rows
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                # Format currency columns (lender columns and statistics)
                if col > 2 and col <= len(self.lenders) + 2:  # Lender columns
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '£#,##0'
                elif col > len(self.lenders) + 2:  # Statistics columns
                    if cell.value and isinstance(cell.value, (int, float)):
                        if 'Rank' in df.columns[col-1]:
                            cell.number_format = '0'
                        else:
                            cell.number_format = '£#,##0'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def manual_data_entry(self) -> None:
        """Interactive function for manual data entry (Phase 1)."""
        print("=== Affordability Tool - Manual Data Entry ===\n")
        
        for scenario_type in ['vanilla', 'self_employed']:
            print(f"\n--- {scenario_type.upper()} SCENARIOS ---")
            
            for scenario in self.scenarios[scenario_type]:
                print(f"\nScenario: {scenario['id']}")
                print(f"Type: {scenario['applicant_type']}")
                print(f"Income: £{scenario['applicant1_income']:,}", end="")
                if scenario['applicant2_income']:
                    print(f" + £{scenario['applicant2_income']:,}")
                else:
                    print()
                print(f"Employment: {scenario['applicant1_employment']}", end="")
                if scenario['applicant2_employment']:
                    print(f" + {scenario['applicant2_employment']}")
                else:
                    print()
                print(f"Notes: {scenario['notes']}")
                
                amounts = {}
                print("\nEnter borrowing amounts for each lender (press Enter to skip):")
                
                for lender in self.lenders:
                    while True:
                        try:
                            amount_str = input(f"{lender}: £").strip()
                            if amount_str == "":
                                amounts[lender] = 0
                                break
                            amount = float(amount_str.replace(',', ''))
                            amounts[lender] = amount
                            break
                        except ValueError:
                            print("Please enter a valid number or press Enter to skip")
                
                self.input_borrowing_amounts(scenario_type, scenario['id'], amounts)
                
                # Show quick stats
                stats = self.calculate_statistics(amounts)
                print(f"\nQuick Stats - Average: £{stats['average']:,.0f}, "
                      f"Gen H Diff: £{stats['gen_h_diff']:,.0f}, "
                      f"Gen H Rank: {stats['gen_h_rank']}")
                
                continue_input = input("\nContinue to next scenario? (y/n): ").lower()
                if continue_input == 'n':
                    return


def main():
    """Main function to run the affordability tool."""
    tool = AffordabilityTool()
    
    print("Affordability Benchmarking Tool")
    print("1. Manual data entry")
    print("2. Load sample data and export")
    print("3. Export current data")
    
    choice = input("\nSelect option (1-3): ").strip()
    
    if choice == "1":
        tool.manual_data_entry()
        export_choice = input("\nExport to Excel? (y/n): ").lower()
        if export_choice == 'y':
            filename = tool.export_to_excel()
            print(f"Results exported to: {filename}")
    
    elif choice == "2":
        # Load sample data for demonstration
        print("Loading sample data...")
        
        # Sample data for vanilla scenarios
        sample_amounts = {
            'Gen H': 180000, 'Accord': 175000, 'Skipton': 182000,
            'Kensington': 170000, 'Precise': 165000, 'Atom': 185000,
            'Clydesdale': 178000, 'Newcastle': 176000, 'Metro': 183000,
            'Nottingham': 174000, 'Hinckley & Rugby': 172000,
            'Leeds': 179000, 'Principality': 181000, 'Coventry': 177000,
            'Santander': 180000
        }
        
        # Add sample data for first few scenarios
        for scenario in tool.scenarios['vanilla'][:2]:
            tool.input_borrowing_amounts('vanilla', scenario['id'], sample_amounts)
        
        for scenario in tool.scenarios['self_employed'][:2]:
            # Slightly different amounts for self-employed
            se_amounts = {k: v * 0.9 for k, v in sample_amounts.items()}
            tool.input_borrowing_amounts('self_employed', scenario['id'], se_amounts)
        
        filename = tool.export_to_excel()
        print(f"Sample results exported to: {filename}")
    
    elif choice == "3":
        filename = tool.export_to_excel()
        print(f"Results exported to: {filename}")
    
    else:
        print("Invalid choice")


if __name__ == "__main__":
    main()